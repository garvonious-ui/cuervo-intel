"""
Excel Dashboard Generator.
Creates a multi-tab Excel report with charts and formatted tables.
"""

import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from templates import BRANDS


# ─── STYLE CONSTANTS ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
CUERVO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HIGH_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
MED_FILL = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
LOW_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B2A4A")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1B2A4A")


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_cuervo=False):
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if is_cuervo:
        cell.fill = CUERVO_FILL


def auto_width(ws, min_width=10, max_width=40):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ─── TAB GENERATORS ──────────────────────────────────────────────────


def create_executive_summary(wb: Workbook, results: dict):
    """Tab 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1B2A4A"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "COMPETITIVE SOCIAL MEDIA INTELLIGENCE REPORT"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1B2A4A")
    ws["A2"] = f"Jose Cuervo vs. Competitors | Generated {datetime.now().strftime('%B %d, %Y')}"
    ws["A2"].font = Font(name="Calibri", size=12, color="666666", italic=True)
    ws.merge_cells("A2:H2")

    row = 4
    ws.cell(row=row, column=1, value="KEY FINDINGS").font = SUBTITLE_FONT
    row += 1

    # Summary stats
    recs = results["recommendations"]
    high_priority = [r for r in recs if r["priority"] == "High"]
    med_priority = [r for r in recs if r["priority"] == "Medium"]

    ws.cell(row=row, column=1, value=f"Total Recommendations: {len(recs)}")
    ws.cell(row=row, column=1).font = BODY_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"High Priority Actions: {len(high_priority)}")
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, color="CC0000", bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Medium Priority Actions: {len(med_priority)}")
    ws.cell(row=row, column=1).font = BODY_FONT
    row += 2

    # Engagement snapshot
    ws.cell(row=row, column=1, value="ENGAGEMENT RATE SNAPSHOT").font = SUBTITLE_FONT
    row += 1

    headers = ["Brand", "IG Followers", "IG Avg ER%", "IG Posts/30d",
               "TT Followers", "TT Avg ER%", "TT Posts/30d"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        ig = results["engagement"].get(brand, {}).get("Instagram", {})
        tt = results["engagement"].get(brand, {}).get("TikTok", {})
        ig_freq = results["frequency"].get(brand, {}).get("Instagram", {})
        tt_freq = results["frequency"].get(brand, {}).get("TikTok", {})

        values = [
            brand,
            ig.get("followers", 0),
            ig.get("avg_engagement_rate", 0),
            ig_freq.get("total_posts_30d", 0),
            tt.get("followers", 0),
            tt.get("avg_engagement_rate", 0),
            tt_freq.get("total_posts_30d", 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    row += 2

    # Top recommendations
    ws.cell(row=row, column=1, value="TOP PRIORITY RECOMMENDATIONS FOR CUERVO").font = SUBTITLE_FONT
    row += 1

    rec_headers = ["#", "Category", "Platform", "Priority", "Insight", "Recommended Action"]
    for c, h in enumerate(rec_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(rec_headers))
    row += 1

    for i, rec in enumerate(recs[:15], 1):
        values = [i, rec["category"], rec["platform"], rec["priority"],
                  rec["insight"], rec["recommendation"]]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c)
            # Color-code priority
            if c == 4:
                cell = ws.cell(row=row, column=c)
                if v == "High":
                    cell.fill = HIGH_FILL
                elif v == "Medium":
                    cell.fill = MED_FILL
                else:
                    cell.fill = LOW_FILL
        row += 1

    auto_width(ws, max_width=60)
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 50


def create_brand_comparison(wb: Workbook, results: dict):
    """Tab 2: Brand-by-Brand Comparison Table."""
    ws = wb.create_sheet("Brand Comparison")
    ws.sheet_properties.tabColor = "4A86C8"

    ws["A1"] = "BRAND-BY-BRAND COMPARISON"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    row = 3
    # Instagram section
    ws.cell(row=row, column=1, value="INSTAGRAM").font = SUBTITLE_FONT
    row += 1

    ig_headers = [
        "Brand", "Followers", "Posts/30d", "Posts/Week",
        "Avg ER%", "Avg Likes", "Avg Comments", "Avg Views",
        "Top Content Type", "Collab %", "Avg Hashtags/Post",
    ]
    for c, h in enumerate(ig_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(ig_headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        eng = results["engagement"].get(brand, {}).get("Instagram", {})
        freq = results["frequency"].get(brand, {}).get("Instagram", {})
        tags = results["hashtags"].get(brand, {})
        creators = results["creators"].get(brand, {})

        # Find top content type
        by_type = freq.get("by_content_type", {})
        top_type = max(by_type, key=by_type.get) if by_type else "N/A"

        values = [
            brand,
            eng.get("followers", 0),
            freq.get("total_posts_30d", 0),
            freq.get("posts_per_week", 0),
            eng.get("avg_engagement_rate", 0),
            eng.get("avg_likes", 0),
            eng.get("avg_comments", 0),
            eng.get("avg_views", 0),
            top_type,
            creators.get("collab_pct", 0),
            tags.get("avg_hashtags_per_post", 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    row += 2

    # TikTok section
    ws.cell(row=row, column=1, value="TIKTOK").font = SUBTITLE_FONT
    row += 1

    tt_headers = [
        "Brand", "Followers", "Posts/30d", "Posts/Week",
        "Avg ER%", "Avg Likes", "Avg Comments", "Avg Shares",
        "Avg Views", "Collab %", "Avg Hashtags/Post",
    ]
    for c, h in enumerate(tt_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(tt_headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        eng = results["engagement"].get(brand, {}).get("TikTok", {})
        freq = results["frequency"].get(brand, {}).get("TikTok", {})
        tags = results["hashtags"].get(brand, {})
        creators = results["creators"].get(brand, {})

        values = [
            brand,
            eng.get("followers", 0),
            freq.get("total_posts_30d", 0),
            freq.get("posts_per_week", 0),
            eng.get("avg_engagement_rate", 0),
            eng.get("avg_likes", 0),
            eng.get("avg_comments", 0),
            eng.get("avg_shares", 0),
            eng.get("avg_views", 0),
            creators.get("collab_pct", 0),
            tags.get("avg_hashtags_per_post", 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    auto_width(ws)


def create_content_strategy(wb: Workbook, results: dict):
    """Tab 3: Content Strategy Insights."""
    ws = wb.create_sheet("Content Strategy")
    ws.sheet_properties.tabColor = "6AA84F"

    ws["A1"] = "CONTENT STRATEGY INSIGHTS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    row = 3

    # Content theme performance matrix
    ws.cell(row=row, column=1, value="CONTENT THEME PERFORMANCE BY BRAND").font = SUBTITLE_FONT
    row += 1

    headers = ["Brand", "Top Theme (Volume)", "Best Theme (ER)", "Best ER%",
               "Top Visual Style", "Avg Caption Words", "Top CTA"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        theme = results["themes"].get(brand, {})
        cap_ig = results["captions"].get(brand, {}).get("Instagram", {})
        cap_tt = results["captions"].get(brand, {}).get("TikTok", {})

        top_theme = theme.get("top_theme", ("N/A", 0))
        best_theme = theme.get("best_performing_theme", "N/A")
        best_er = 0
        if best_theme != "N/A":
            perf = theme.get("theme_performance", {}).get(best_theme, {})
            best_er = perf.get("avg_engagement_rate", 0)

        styles = theme.get("visual_style_distribution", {})
        top_style = max(styles, key=styles.get) if styles else "N/A"

        avg_words = round((cap_ig.get("avg_word_count", 0) + cap_tt.get("avg_word_count", 0)) / 2, 1)

        # Combine CTAs across platforms
        all_ctas = {}
        for cap in [cap_ig, cap_tt]:
            for cta, count in cap.get("top_ctas", []):
                all_ctas[cta] = all_ctas.get(cta, 0) + count
        top_cta = max(all_ctas, key=all_ctas.get) if all_ctas else "None"

        values = [
            brand,
            f"{top_theme[0]} ({top_theme[1]} posts)" if isinstance(top_theme, tuple) else str(top_theme),
            best_theme,
            best_er,
            top_style,
            avg_words,
            top_cta,
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    row += 2

    # Detailed theme breakdown per brand
    ws.cell(row=row, column=1, value="DETAILED THEME BREAKDOWN").font = SUBTITLE_FONT
    row += 1

    detail_headers = ["Brand", "Theme", "Post Count", "% of Content", "Avg ER%"]
    for c, h in enumerate(detail_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(detail_headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        theme_perf = results["themes"].get(brand, {}).get("theme_performance", {})

        # Sort by post count
        sorted_themes = sorted(theme_perf.items(), key=lambda x: x[1]["count"], reverse=True)
        for theme_name, perf in sorted_themes:
            values = [
                brand, theme_name, perf["count"],
                f"{perf['pct_of_content']}%", perf["avg_engagement_rate"],
            ]
            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1

    row += 2

    # Caption analysis
    ws.cell(row=row, column=1, value="CAPTION ANALYSIS").font = SUBTITLE_FONT
    row += 1

    cap_headers = ["Brand", "Platform", "Avg Words", "Avg Emojis",
                   "Dominant Tone", "Length vs. Engagement"]
    for c, h in enumerate(cap_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(cap_headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        for platform in ["Instagram", "TikTok"]:
            cap = results["captions"].get(brand, {}).get(platform, {})
            tones = cap.get("tone_distribution", {})
            top_tone = max(tones, key=tones.get) if tones else "N/A"

            values = [
                brand, platform,
                cap.get("avg_word_count", 0),
                cap.get("avg_emoji_count", 0),
                top_tone,
                cap.get("caption_length_vs_engagement", "N/A"),
            ]
            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1

    auto_width(ws, max_width=45)


def create_engagement_benchmarks(wb: Workbook, results: dict):
    """Tab 4: Engagement Benchmarks."""
    ws = wb.create_sheet("Engagement Benchmarks")
    ws.sheet_properties.tabColor = "E69138"

    ws["A1"] = "ENGAGEMENT BENCHMARKS BY CONTENT TYPE"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")

    row = 3

    for platform in ["Instagram", "TikTok"]:
        ws.cell(row=row, column=1, value=f"{platform.upper()} ENGAGEMENT RATES BY CONTENT TYPE").font = SUBTITLE_FONT
        row += 1

        # Collect all content types across brands
        all_types = set()
        for brand in BRANDS:
            eng = results["engagement"].get(brand, {}).get(platform, {})
            all_types.update(eng.get("engagement_by_type", {}).keys())
        all_types = sorted(all_types)

        headers = ["Brand"] + list(all_types) + ["Overall Avg ER%"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        # Chart data start
        chart_start_row = row

        for brand in BRANDS:
            is_cuervo = brand == "Jose Cuervo"
            eng = results["engagement"].get(brand, {}).get(platform, {})
            by_type = eng.get("engagement_by_type", {})

            values = [brand]
            for t in all_types:
                values.append(by_type.get(t, 0))
            values.append(eng.get("avg_engagement_rate", 0))

            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1

        # Add a bar chart
        if all_types and len(BRANDS) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = f"{platform} Avg Engagement Rate by Content Type"
            chart.y_axis.title = "Engagement Rate %"
            chart.x_axis.title = "Brand"
            chart.style = 10
            chart.width = 22
            chart.height = 12

            # Overall ER column
            overall_col = len(headers)
            data = Reference(ws, min_col=overall_col, min_row=chart_start_row - 1,
                             max_row=chart_start_row + len(BRANDS) - 1)
            cats = Reference(ws, min_col=1, min_row=chart_start_row,
                             max_row=chart_start_row + len(BRANDS) - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, f"A{row + 1}")
            row += 17  # Space for chart

        row += 2

    # Creator collab engagement comparison
    ws.cell(row=row, column=1, value="CREATOR COLLAB VS. BRAND CONTENT ENGAGEMENT").font = SUBTITLE_FONT
    row += 1

    collab_headers = ["Brand", "Collab Posts", "Collab ER%", "Non-Collab ER%",
                      "ER Lift from Collabs", "Unique Creators", "Paid Partnerships"]
    for c, h in enumerate(collab_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(collab_headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        cr = results["creators"].get(brand, {})
        values = [
            brand,
            cr.get("total_collab_posts", 0),
            cr.get("avg_collab_engagement_rate", 0),
            cr.get("avg_non_collab_engagement_rate", 0),
            cr.get("collab_engagement_lift", 0),
            cr.get("unique_creators", 0),
            cr.get("paid_partnerships", 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    auto_width(ws)


def create_hashtag_analysis(wb: Workbook, results: dict):
    """Tab 5: Hashtag Strategy Comparison."""
    ws = wb.create_sheet("Hashtag Strategy")
    ws.sheet_properties.tabColor = "A64D79"

    ws["A1"] = "HASHTAG STRATEGY COMPARISON"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    row = 3

    # Overview
    ws.cell(row=row, column=1, value="HASHTAG USAGE OVERVIEW").font = SUBTITLE_FONT
    row += 1

    headers = ["Brand", "Unique Hashtags", "Avg per Post", "Branded %",
               "Total Tag Usage (30d)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        tags = results["hashtags"].get(brand, {})
        values = [
            brand,
            tags.get("unique_hashtags", 0),
            tags.get("avg_hashtags_per_post", 0),
            f"{tags.get('branded_hashtag_pct', 0)}%",
            tags.get("total_hashtag_usage", 0),
        ]
        for c, v in enumerate(values, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_cuervo)
        row += 1

    row += 2

    # Top hashtags per brand
    ws.cell(row=row, column=1, value="TOP 15 HASHTAGS PER BRAND").font = SUBTITLE_FONT
    row += 1

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        ws.cell(row=row, column=1, value=brand).font = SUBHEADER_FONT
        if is_cuervo:
            ws.cell(row=row, column=1).fill = CUERVO_FILL
        row += 1

        tag_headers = ["Rank", "Hashtag", "Times Used"]
        for c, h in enumerate(tag_headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(tag_headers))
        row += 1

        tags = results["hashtags"].get(brand, {})
        for i, (tag, count) in enumerate(tags.get("top_15_hashtags", []), 1):
            values = [i, tag, count]
            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1
        row += 1

    auto_width(ws)


def create_top_posts(wb: Workbook, results: dict):
    """Tab 6: Top Performing Content Examples."""
    ws = wb.create_sheet("Top Posts")
    ws.sheet_properties.tabColor = "3D85C6"

    ws["A1"] = "TOP PERFORMING CONTENT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")

    row = 3

    for brand in BRANDS:
        is_cuervo = brand == "Jose Cuervo"
        ws.cell(row=row, column=1, value=brand.upper()).font = SUBTITLE_FONT
        if is_cuervo:
            ws.cell(row=row, column=1).fill = CUERVO_FILL
        row += 1

        for platform in ["Instagram", "TikTok"]:
            ws.cell(row=row, column=1, value=platform).font = SUBHEADER_FONT
            row += 1

            top_headers = ["Rank", "URL", "Date", "Type", "ER%",
                           "Likes", "Comments", "Views", "Theme", "Caption Preview"]
            for c, h in enumerate(top_headers, 1):
                ws.cell(row=row, column=c, value=h)
            style_header_row(ws, row, len(top_headers))
            row += 1

            eng = results["engagement"].get(brand, {}).get(platform, {})
            top_posts = eng.get("top_10_posts", [])

            if not top_posts:
                ws.cell(row=row, column=1, value="No data collected")
                ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True, color="999999")
                row += 1
            else:
                for i, post in enumerate(top_posts, 1):
                    values = [
                        i,
                        post.get("url", ""),
                        post.get("date", ""),
                        post.get("type", ""),
                        post.get("engagement_rate", 0),
                        post.get("likes", 0),
                        post.get("comments", 0),
                        post.get("views", 0),
                        post.get("theme", ""),
                        post.get("caption_preview", ""),
                    ]
                    for c, v in enumerate(values, 1):
                        ws.cell(row=row, column=c, value=v)
                        style_data_cell(ws, row, c, is_cuervo)
                    row += 1

            row += 1
        row += 1

    auto_width(ws, max_width=50)
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["J"].width = 45


def create_posting_schedule(wb: Workbook, results: dict):
    """Tab 7: Posting Schedule Heatmap Data."""
    ws = wb.create_sheet("Posting Schedule")
    ws.sheet_properties.tabColor = "674EA7"

    ws["A1"] = "POSTING SCHEDULE ANALYSIS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    row = 3

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for platform in ["Instagram", "TikTok"]:
        ws.cell(row=row, column=1, value=f"{platform.upper()} - POSTS BY DAY OF WEEK").font = SUBTITLE_FONT
        row += 1

        headers = ["Brand"] + days_order + ["Total"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for brand in BRANDS:
            is_cuervo = brand == "Jose Cuervo"
            freq = results["frequency"].get(brand, {}).get(platform, {})
            by_day = freq.get("by_day", {})

            values = [brand]
            total = 0
            for day in days_order:
                count = by_day.get(day, 0)
                values.append(count)
                total += count
            values.append(total)

            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1

        row += 2

        # Best hours
        ws.cell(row=row, column=1, value=f"{platform.upper()} - BEST POSTING HOURS").font = SUBTITLE_FONT
        row += 1

        hour_headers = ["Brand", "Best Hour #1", "Best Hour #2", "Best Hour #3"]
        for c, h in enumerate(hour_headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(hour_headers))
        row += 1

        for brand in BRANDS:
            is_cuervo = brand == "Jose Cuervo"
            freq = results["frequency"].get(brand, {}).get(platform, {})
            best_hours = freq.get("best_hours", [])

            values = [brand]
            for i in range(3):
                if i < len(best_hours):
                    h, count = best_hours[i]
                    values.append(f"{h}:00 ({count} posts)")
                else:
                    values.append("N/A")

            for c, v in enumerate(values, 1):
                ws.cell(row=row, column=c, value=v)
                style_data_cell(ws, row, c, is_cuervo)
            row += 1

        row += 3

    auto_width(ws)


def create_raw_data(wb: Workbook, posts: list[dict]):
    """Tab 8: Raw Post Data for reference."""
    ws = wb.create_sheet("Raw Data")
    ws.sheet_properties.tabColor = "999999"

    headers = [
        "brand", "platform", "post_url", "post_date", "post_time",
        "post_type", "video_length_seconds", "likes", "comments",
        "shares", "saves", "views", "total_engagement", "engagement_rate",
        "content_theme", "visual_style", "caption_tone", "cta_type",
        "has_creator_collab", "caption_word_count", "emoji_count_in_caption",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for r, post in enumerate(posts, 2):
        for c, h in enumerate(headers, 1):
            val = post.get(h, "")
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).font = BODY_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER

    auto_width(ws)


def generate_dashboard(results: dict, output_path: str) -> str:
    """Generate the complete Excel dashboard."""
    wb = Workbook()

    create_executive_summary(wb, results)
    create_brand_comparison(wb, results)
    create_content_strategy(wb, results)
    create_engagement_benchmarks(wb, results)
    create_hashtag_analysis(wb, results)
    create_top_posts(wb, results)
    create_posting_schedule(wb, results)
    create_raw_data(wb, results.get("posts", []))

    wb.save(output_path)
    return output_path
